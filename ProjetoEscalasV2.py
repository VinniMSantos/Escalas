import sys
import json
import locale
import datetime
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import win32com.client  # Para interação com Outlook

from PyQt5.QtWidgets import (
    QApplication, QWidget, QHBoxLayout, QVBoxLayout, QLabel, QComboBox, QDateTimeEdit,
    QPushButton, QLineEdit, QCompleter, QTableWidget, QTableWidgetItem, QHeaderView,
    QSizePolicy, QFileDialog, QDialog, QDateEdit, QMessageBox,
    QAbstractItemView, QRadioButton, QButtonGroup, QScrollArea, QFormLayout, QCheckBox,
    QMenu, QAction
)
from PyQt5.QtCore import (
    Qt, QDateTime, QDate, QTime, QStringListModel, QTimer, QSortFilterProxyModel,
    QRegularExpression, QPoint, pyqtSignal, QEvent
)
from PyQt5.QtGui import QFont, QIcon, QMouseEvent, QColor, QBrush, QStandardItemModel, QStandardItem

# --------------------------------------------------
#                DICIONÁRIOS DE E-MAIL
# --------------------------------------------------
unit_manager_emails = {
    "HM Benedicto": "gessica.neves@libertyti.com.br",
    "HM Campo Limpo": "eduardo.lima@libertyti.com.br",
    "HM Tatuape": "vinicius.santos@libertyti.com.br",
    "HM Tide": "natalia.lima@libertyti.com.br",
    # ... (se houver mais unidades, pode acrescentar aqui)
}

technician_emails = {
    "Allef Barbosa": "vinicius.santos@libertyti.com.br",
    "Eduardo Lima": "vinicius.santos@libertyti.com.br",
    "Kaue Rodrigues": "kaue.rodrigues@libertyti.com.br",
    "Geovanna Oliveira": "vinicius.santos@libertyti.com.br",
    "Gustavo Silva": "gustavo.silva@libertyti.com.br",
    "Vitor Martins": "vitor.martins@libertyti.com.br",
    "Mateus Marinho": "vinicius.santos@libertyti.com.br",
    "Joao Marinho": "vinicius.santos@libertyti.com.br",
    "Andre Assis": "vinicius.santos@libertyti.com.br",
    "Valdemir Araujo": "vinicius.santos@libertyti.com.br"
}

# --------------------------------------------------
#               FUNÇÃO PARA ENVIAR E-MAIL
# --------------------------------------------------
def send_email(to_email, subject, html_body, send_time=None):
    try:
        outlook = win32com.client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to_email
        mail.Subject = subject
        mail.HTMLBody = html_body

        # Ajustar caminho da assinatura, se necessário
        assinatura_path = r"C:\Users\LIBERTY\Documents\escalas versao\Escalas\assinatura_vinicius.png.png"
        attachment = mail.Attachments.Add(assinatura_path)
        attachment.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
            "MinhaImagem"
        )

        if send_time:
            if send_time <= datetime.datetime.now():
                pass  # Envia imediatamente
            else:
                mail.DeferredDeliveryTime = send_time

        mail.Send()
        namespace = outlook.GetNamespace("MAPI")
        namespace.SendAndReceive(False)

    except Exception as e:
        QMessageBox.warning(None, "Erro", f"Falha ao enviar e-mail para {to_email}: {e}")

# --------------------------------------------------
#               CLASSE LABEL CLICÁVEL
# --------------------------------------------------
class ClickableLabel(QLabel):
    clicked = pyqtSignal()
    def mousePressEvent(self, event):
        self.clicked.emit()

# --------------------------------------------------
#     CLASSE PARA FILTRO DE SUBSTRING NO QComboBox
#       (USADO SOMENTE PARA 'UNIDADE' NESTE EXEMPLO)
# --------------------------------------------------
class SubstringFilterProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.setFilterKeyColumn(0)

    def filterAcceptsRow(self, source_row, source_parent):
        if not self.filterRegularExpression().pattern():
            return True
        model = self.sourceModel()
        index = model.index(source_row, 0, source_parent)
        data = model.data(index, Qt.DisplayRole)
        if data and self.filterRegularExpression().match(data).hasMatch():
            return True
        return False

# --------------------------------------------------
#       COMBOBOX COM FILTRO DE SUBSTRING
#         (Apenas para UNIDADE neste exemplo)
# --------------------------------------------------
class FilteredComboBox(QComboBox):
    """ComboBox filtrado por substring."""
    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)

        self.model_base = QStringListModel(items)
        self.proxy_model = SubstringFilterProxyModel(self)
        self.proxy_model.setSourceModel(self.model_base)

        self.completer = QCompleter(self.proxy_model, self)
        self.completer.setCompletionMode(QCompleter.PopupCompletion)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.setCompleter(self.completer)

        self.lineEdit().textEdited.connect(self.filter_items)
        self.lineEdit().installEventFilter(self)

        self.setStyleSheet("""
            QComboBox {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 6px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
            QComboBox QAbstractItemView {
                background-color: #fff;
                selection-background-color: #007bff;
                selection-color: #fff;
            }
            QComboBox::down-arrow {
                image: url('icons/down-arrow.png');
                width: 14px;
                height: 14px;
            }
        """)

    def filter_items(self, text):
        pattern = f".*{QRegularExpression.escape(text)}.*"
        regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
        self.proxy_model.setFilterRegularExpression(regex)
        self.completer.complete()

    def showPopup(self):
        self.filter_items(self.lineEdit().text())
        super().showPopup()

    def eventFilter(self, source, event):
        if event.type() == QEvent.KeyPress and source is self.lineEdit():
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                pass
        return super().eventFilter(source, event)

# --------------------------------------------------
#   CLASSE PARA MULTI-SELEÇÃO DE TÉCNICOS
#        (PARA O FILTRO DE CONSULTA)
# --------------------------------------------------
class MultiSelectComboBox(QComboBox):
    """
    Exibe um menu de checkboxes para cada item da lista.
    Permite selecionar vários itens, retornando em `checkedItems()`.
    """
    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.items = items
        self.checked_items = []
        self.setEditable(True)  # Para manter tamanho similar ao FilteredComboBox
        self.setInsertPolicy(QComboBox.NoInsert)
        self.setStyleSheet("""
            QComboBox {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
        """)
        # Model para exibir o texto "vazio" no QComboBox (porque usaremos menu)
        self.model_base = QStringListModel(items)
        self.setMaxVisibleItems(len(items))  # Ajusta dropdown, se quiser

    def showPopup(self):
        menu = QMenu(self)
        self.actions = []

        for item in self.items:
            action = QAction(item, menu)
            action.setCheckable(True)
            action.setChecked(item in self.checked_items)
            action.triggered.connect(self.update_checked_items)
            self.actions.append(action)
            menu.addAction(action)

        menu.popup(self.mapToGlobal(self.rect().bottomLeft()))

    def update_checked_items(self):
        self.checked_items = [action.text() for action in self.actions if action.isChecked()]
        # Atualiza texto do ComboBox
        if self.checked_items:
            display_text = ", ".join(self.checked_items)
        else:
            display_text = ""
        self.setCurrentText(display_text)

    def checkedItems(self):
        """Retorna a lista de itens selecionados."""
        return self.checked_items

# --------------------------------------------------
#  CLASSE DE SELEÇÃO DE PLANILHA E PERÍODO INICIAL
# --------------------------------------------------
class SelectionDialog(QDialog):
    """
    Tela inicial que pergunta Planilha + Período.
    Esse período será usado apenas se escolher "Incluir Escala".
    Caso seja "Consultar Escala", será ignorado.
    """
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        self.planilha_path = None
        self.periodo_inicio = None
        self.periodo_fim = None
        self.choice = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Seleção de Planilha e Período")
        self.resize(400, 300)
        layout = QVBoxLayout()

        self.planilha_label = QLabel("Nenhuma planilha selecionada")
        self.planilha_label.setAlignment(Qt.AlignCenter)
        self.planilha_button = QPushButton("Selecionar Planilha")
        self.planilha_button.clicked.connect(self.select_planilha)
        self.planilha_button.setStyleSheet(self.get_button_style())

        periodo_layout = QHBoxLayout()
        self.data_inicio = QDateEdit()
        self.data_inicio.setCalendarPopup(True)
        self.data_inicio.setDate(QDate.currentDate())

        self.data_fim = QDateEdit()
        self.data_fim.setCalendarPopup(True)
        self.data_fim.setDate(QDate.currentDate())

        periodo_layout.addWidget(QLabel("Data Início:"))
        periodo_layout.addWidget(self.data_inicio)
        periodo_layout.addWidget(QLabel("Data Fim:"))
        periodo_layout.addWidget(self.data_fim)

        self.radio_incluir_escala = QRadioButton("Incluir Escala")
        self.radio_consultar_escala = QRadioButton("Consultar Escala")
        self.radio_incluir_escala.setChecked(True)

        self.choice_group = QButtonGroup()
        self.choice_group.addButton(self.radio_incluir_escala)
        self.choice_group.addButton(self.radio_consultar_escala)

        choice_layout = QHBoxLayout()
        choice_layout.addWidget(self.radio_incluir_escala)
        choice_layout.addWidget(self.radio_consultar_escala)

        self.confirm_button = QPushButton("Confirmar")
        self.confirm_button.clicked.connect(self.confirm_selection)
        self.confirm_button.setEnabled(False)
        self.confirm_button.setStyleSheet(self.get_button_style())

        layout.addWidget(self.planilha_label)
        layout.addWidget(self.planilha_button)
        layout.addLayout(periodo_layout)
        layout.addLayout(choice_layout)
        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def select_planilha(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Planilha", "", "Planilhas Excel (*.xlsx *.xls)", options=options
        )
        if file_path:
            self.planilha_path = file_path
            self.planilha_label.setText(f"Planilha selecionada:\n{file_path}")
            self.confirm_button.setEnabled(True)

    def confirm_selection(self):
        self.periodo_inicio = self.data_inicio.date()
        self.periodo_fim = self.data_fim.date()
        if self.periodo_fim < self.periodo_inicio:
            QMessageBox.warning(self, "Erro", "Data fim não pode ser antes da data início.")
        else:
            if self.radio_incluir_escala.isChecked():
                self.choice = 'incluir_escala'
            elif self.radio_consultar_escala.isChecked():
                self.choice = 'consultar_escala'
            self.accept()

    def get_button_style(self):
        return """
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:disabled {
                background-color: #6c757d;
            }
        """

# --------------------------------------------------
#     CLASSE PARA ESCOLHER PERÍODO DE CONSULTA
# --------------------------------------------------
class PeriodoConsultaDialog(QDialog):
    """
    Caso o usuário queira alterar o período
    enquanto está em "Incluir Escala".
    """
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        self.periodo_inicio = None
        self.periodo_fim = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Período de Consulta")
        self.resize(300, 150)
        layout = QVBoxLayout()

        periodo_layout = QHBoxLayout()
        self.data_inicio = QDateEdit()
        self.data_inicio.setCalendarPopup(True)
        self.data_inicio.setDate(QDate.currentDate())
        self.data_fim = QDateEdit()
        self.data_fim.setCalendarPopup(True)
        self.data_fim.setDate(QDate.currentDate())

        periodo_layout.addWidget(QLabel("Data Início:"))
        periodo_layout.addWidget(self.data_inicio)
        periodo_layout.addWidget(QLabel("Data Fim:"))
        periodo_layout.addWidget(self.data_fim)

        self.confirm_button = QPushButton("Confirmar")
        self.confirm_button.clicked.connect(self.confirm_selection)
        self.confirm_button.setStyleSheet(self.get_button_style())

        layout.addLayout(periodo_layout)
        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def confirm_selection(self):
        self.periodo_inicio = self.data_inicio.date()
        self.periodo_fim = self.data_fim.date()
        if self.periodo_fim < self.periodo_inicio:
            QMessageBox.warning(self, "Erro", "Data fim não pode ser antes da data início.")
        else:
            self.accept()

    def get_button_style(self):
        return """
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """

# --------------------------------------------------
#   CLASSE PARA SELEÇÃO DE TÉCNICOS (12x36, 5x2)
# --------------------------------------------------
class TechnicianSelectionDialog(QDialog):
    """Seleciona múltiplos técnicos, definindo 'pares' ou 'ímpares' se 12x36."""
    def __init__(self, tecnicos, technician_schedules):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        self.selected_tecnicos = {}
        self.technician_schedules = technician_schedules
        self.init_ui(tecnicos)

    def init_ui(self, tecnicos):
        self.setWindowTitle("Selecionar Técnicos")
        self.resize(400, 500)
        layout = QVBoxLayout()

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        widget = QWidget()
        self.scroll_area.setWidget(widget)
        form_layout = QFormLayout(widget)

        self.tecnico_widgets = {}

        for tecnico in tecnicos:
            tecnico_info = self.technician_schedules.get(tecnico, {})
            escala = tecnico_info.get('escala')
            hbox = QHBoxLayout()
            checkbox = QCheckBox()
            hbox.addWidget(checkbox)
            label = QLabel(tecnico)
            hbox.addWidget(label)

            if escala == '12X36':
                radiobutton_pares = QRadioButton("Pares")
                radiobutton_impares = QRadioButton("Ímpares")
                button_group = QButtonGroup(self)
                button_group.addButton(radiobutton_pares)
                button_group.addButton(radiobutton_impares)
                hbox.addWidget(radiobutton_pares)
                hbox.addWidget(radiobutton_impares)
                self.tecnico_widgets[tecnico] = (checkbox, radiobutton_pares, radiobutton_impares)
            else:
                hbox.addStretch()
                self.tecnico_widgets[tecnico] = (checkbox, None, None)

            form_layout.addRow(hbox)

        self.confirm_button = QPushButton("Confirmar")
        self.confirm_button.clicked.connect(self.confirm_selection)
        self.confirm_button.setStyleSheet(self.get_button_style())

        layout.addWidget(self.scroll_area)
        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def confirm_selection(self):
        self.selected_tecnicos = {}
        for tecnico, widgets in self.tecnico_widgets.items():
            checkbox, radiobutton_pares, radiobutton_impares = widgets
            if checkbox.isChecked():
                if radiobutton_pares and radiobutton_impares:
                    if radiobutton_pares.isChecked():
                        dias_trabalho = 'pares'
                    elif radiobutton_impares.isChecked():
                        dias_trabalho = 'impares'
                    else:
                        QMessageBox.warning(
                            self,
                            "Aviso",
                            f"Selecione 'Pares' ou 'Ímpares' para o técnico {tecnico}."
                        )
                        return
                    self.selected_tecnicos[tecnico] = dias_trabalho
                else:
                    self.selected_tecnicos[tecnico] = None
        if not self.selected_tecnicos:
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um técnico.")
        else:
            self.accept()

    def get_button_style(self):
        return """
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """

# --------------------------------------------------
#   CLASSE PARA ESCOLHER TÉCNICO (via substring) +
#        PERÍODO PARA ENVIO DE E-MAIL
# --------------------------------------------------
class EmailSelectionDialog(QDialog):
    """Permite digitar parte do nome do técnico e escolher período para o e-mail."""
    def __init__(self, tecnicos, periodo_inicio, periodo_fim):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        self.selected_tecnicos = []
        self.periodo_inicio = periodo_inicio
        self.periodo_fim = periodo_fim
        self.init_ui(tecnicos)

    def init_ui(self, tecnicos):
        self.setWindowTitle("Selecionar Técnicos e Período para Envio")
        self.resize(400, 500)
        layout = QVBoxLayout()

        self.tecnico_line_edit = QLineEdit()
        self.tecnico_line_edit.setPlaceholderText("Digite (parte do) nome do técnico")
        layout.addWidget(QLabel("Filtrar Técnico:"))
        layout.addWidget(self.tecnico_line_edit)

        periodo_layout = QHBoxLayout()
        self.data_inicio = QDateEdit()
        self.data_inicio.setCalendarPopup(True)
        self.data_inicio.setDate(self.periodo_inicio)
        self.data_fim = QDateEdit()
        self.data_fim.setCalendarPopup(True)
        self.data_fim.setDate(self.periodo_fim)

        periodo_layout.addWidget(QLabel("Data Início:"))
        periodo_layout.addWidget(self.data_inicio)
        periodo_layout.addWidget(QLabel("Data Fim:"))
        periodo_layout.addWidget(self.data_fim)

        layout.addWidget(QLabel("Selecione o Período:"))
        layout.addLayout(periodo_layout)

        self.confirm_button = QPushButton("Confirmar")
        self.confirm_button.clicked.connect(self.confirm_selection)
        self.confirm_button.setStyleSheet(self.get_button_style())

        layout.addWidget(self.confirm_button)
        self.setLayout(layout)

    def confirm_selection(self):
        typed_text = self.tecnico_line_edit.text().strip()
        if not typed_text:
            QMessageBox.warning(self, "Aviso", "Digite pelo menos parte do nome de um técnico.")
            return
        self.selected_tecnicos = [typed_text]
        self.periodo_inicio = self.data_inicio.date()
        self.periodo_fim = self.data_fim.date()
        if self.periodo_fim < self.periodo_inicio:
            QMessageBox.warning(self, "Erro", "Data fim não pode ser antes da data início.")
            return
        self.accept()

    def get_button_style(self):
        return """
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                padding: 6px 12px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """

# --------------------------------------------------
#   TELA DE CONSULTA COM FILTRO (TÉCNICO, UNIDADE)
#   E FILTRO DE DATA -- MULTI-SELEÇÃO DE TÉCNICOS
# --------------------------------------------------
class ConsultaEscalaDialog(QDialog):
    """
    Carrega TODOS os dados (df_filtered) e permite
    filtrar por Técnico (agora múltiplos), Unidade e Período (data).
    """
    def __init__(self, df_filtered, planilha_path, df_existing, periodo_inicio, periodo_fim, labels):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        # Este df_filtered deve conter todos os dados da planilha.
        self.original_df = df_filtered.copy()
        self.df_filtered = df_filtered.copy()
        self.planilha_path = planilha_path
        self.df_existing = df_existing.copy()
        self.periodo_inicio = periodo_inicio
        self.periodo_fim = periodo_fim
        self.labels = labels

        # NOVO: armazenará SEQ das linhas excluídas
        self.deleted_seq = set()

        self.sort_columns = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Consulta de Escala")
        self.showMaximized()
        layout = QVBoxLayout()

        # FILTRO EM LINHA (Técnico, Unidade, Data Início, Data Fim)
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        tecnico_label = QLabel("Técnico:")

        # Lista de técnicos para multi-seleção
        if 'TÉCNICO' in self.original_df.columns:
            lista_tecnicos = sorted(self.original_df['TÉCNICO'].dropna().unique())
        else:
            lista_tecnicos = []
        # Aqui usamos a MultiSelectComboBox
        self.tecnico_combo = MultiSelectComboBox(lista_tecnicos, parent=self)
        self.tecnico_combo.setFixedWidth(120)

        unidade_label = QLabel("Unidade:")
        if 'UNIDADE' in self.original_df.columns:
            lista_unidades = sorted(self.original_df['UNIDADE'].dropna().unique())
        else:
            lista_unidades = []
        self.unidade_combo = FilteredComboBox(lista_unidades, parent=self)
        self.unidade_combo.setFixedWidth(120)

        # Período de filtro
        data_inicio_label = QLabel("Início:")
        self.data_inicio_filter = QDateEdit()
        self.data_inicio_filter.setCalendarPopup(True)
        self.data_inicio_filter.setDate(self.periodo_inicio)
        self.data_inicio_filter.setFixedWidth(110)

        data_fim_label = QLabel("Fim:")
        self.data_fim_filter = QDateEdit()
        self.data_fim_filter.setCalendarPopup(True)
        self.data_fim_filter.setDate(self.periodo_fim)
        self.data_fim_filter.setFixedWidth(110)

        self.apply_filter_button = QPushButton("Filtrar")
        self.apply_filter_button.clicked.connect(self.apply_filter)
        self.apply_filter_button.setFixedHeight(30)

        self.clear_filter_button = QPushButton("Limpar")
        self.clear_filter_button.clicked.connect(self.clear_filter)
        self.clear_filter_button.setFixedHeight(30)
        self.clear_filter_button.setEnabled(False)

        filter_layout.addWidget(tecnico_label)
        filter_layout.addWidget(self.tecnico_combo)
        filter_layout.addWidget(unidade_label)
        filter_layout.addWidget(self.unidade_combo)
        filter_layout.addWidget(data_inicio_label)
        filter_layout.addWidget(self.data_inicio_filter)
        filter_layout.addWidget(data_fim_label)
        filter_layout.addWidget(self.data_fim_filter)
        filter_layout.addWidget(self.apply_filter_button)
        filter_layout.addWidget(self.clear_filter_button)

        layout.addLayout(filter_layout)

        # Ordena colunas
        self.df_filtered = self.df_filtered[self.labels]

        # Converte data/hora
        for col in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
            self.df_filtered[col] = pd.to_datetime(
                self.df_filtered[col], format="%d/%m/%Y %H:%M:%S", errors='coerce'
            )

        self.table_widget = QTableWidget()
        self.table_widget.setRowCount(len(self.df_filtered))
        self.table_widget.setColumnCount(len(self.df_filtered.columns))
        self.table_widget.setHorizontalHeaderLabels(self.df_filtered.columns.tolist())
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setMinimumSectionSize(100)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
            }
            QTableWidget::item:selected {
                background-color: #007bff;
                color: #fff;
            }
            QHeaderView::section {
                background-color: #343a40;
                color: white;
                font-weight: bold;
                font-size: 14px;
                height: 30px;
            }
        """)

        self.table_widget.horizontalHeader().setSectionsClickable(True)
        self.table_widget.horizontalHeader().setSortIndicatorShown(True)
        self.table_widget.horizontalHeader().sectionClicked.connect(self.handle_header_click)

        self.populate_table()
        layout.addWidget(self.table_widget)

        # Botões: Editar, Excluir, Salvar, Enviar
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        self.edit_button = QPushButton(" Editar")
        self.edit_button.setFixedSize(140, 40)
        self.edit_button.setIcon(QIcon("icons/edit.png"))
        self.edit_button.setStyleSheet(self.get_warning_button_style())
        self.edit_button.clicked.connect(self.enable_editing)
        buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton(" Excluir")
        self.delete_button.setFixedSize(140, 40)
        self.delete_button.setIcon(QIcon("icons/delete.png"))
        self.delete_button.setStyleSheet(self.get_danger_button_style())
        self.delete_button.clicked.connect(self.delete_entry)
        buttons_layout.addWidget(self.delete_button)

        self.save_button = QPushButton(" Salvar Alterações")
        self.save_button.setFixedSize(180, 40)
        self.save_button.setIcon(QIcon("icons/save.png"))
        self.save_button.setStyleSheet(self.get_primary_button_style())
        self.save_button.clicked.connect(self.save_changes)
        buttons_layout.addWidget(self.save_button)

        self.send_email_button = QPushButton(" Enviar Escala")
        self.send_email_button.setFixedSize(180, 40)
        self.send_email_button.setIcon(QIcon("icons/email.png"))
        self.send_email_button.setStyleSheet(self.get_primary_button_style())
        self.send_email_button.clicked.connect(self.send_emails)
        buttons_layout.addWidget(self.send_email_button)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def apply_filter(self):
        """Filtrar por (múltiplos) Técnicos, Unidade e Data (início/fim)."""
        selected_techs = self.tecnico_combo.checkedItems()  # lista de técnicos
        unit_text = self.unidade_combo.currentText().strip()

        df_temp = self.original_df.copy()

        # Se houver técnicos marcados, filtra por esses nomes (OR).
        if selected_techs:
            df_temp = df_temp[df_temp['TÉCNICO'].isin(selected_techs)]

        if unit_text and 'UNIDADE' in df_temp.columns:
            df_temp = df_temp[df_temp['UNIDADE'].str.contains(unit_text, case=False, na=False)]

        start_qdate = self.data_inicio_filter.date()
        end_qdate = self.data_fim_filter.date()
        if end_qdate < start_qdate:
            QMessageBox.warning(self, "Aviso", "Data fim não pode ser antes da data início.")
            return

        start_date = QDateTime(start_qdate, QTime(0, 0, 0)).toPyDateTime()
        end_date = QDateTime(end_qdate, QTime(23, 59, 59)).toPyDateTime()

        df_temp['DATA/HORA INICIO'] = pd.to_datetime(df_temp['DATA/HORA INICIO'], errors='coerce')
        df_temp = df_temp[
            (df_temp['DATA/HORA INICIO'] >= start_date) &
            (df_temp['DATA/HORA INICIO'] <= end_date)
        ]

        self.df_filtered = df_temp.copy()

        for col in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
            self.df_filtered[col] = pd.to_datetime(
                self.df_filtered[col], format="%d/%m/%Y %H:%M:%S", errors='coerce'
            )
        self.df_filtered = self.df_filtered[self.labels]
        self.sort_columns.clear()
        self.populate_table()
        self.clear_filter_button.setEnabled(True)

    def clear_filter(self):
        """Remove o filtro e exibe todos os dados novamente."""
        self.df_filtered = self.original_df.copy()

        for col in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
            self.df_filtered[col] = pd.to_datetime(
                self.df_filtered[col], format="%d/%m/%Y %H:%M:%S", errors='coerce'
            )
        self.df_filtered = self.df_filtered[self.labels]
        self.sort_columns.clear()
        self.populate_table()

        self.clear_filter_button.setEnabled(False)
        # "Limpar" a MultiSelectComboBox
        self.tecnico_combo.checked_items.clear()
        self.tecnico_combo.setCurrentText("")
        self.unidade_combo.setCurrentText("")
        self.data_inicio_filter.setDate(self.periodo_inicio)
        self.data_fim_filter.setDate(self.periodo_fim)

    def populate_table(self):
        self.table_widget.clearContents()
        self.table_widget.setRowCount(len(self.df_filtered))

        for row in range(len(self.df_filtered)):
            for col in range(len(self.df_filtered.columns)):
                value = self.df_filtered.iloc[row, col]
                if pd.isna(value):
                    display_value = ''
                else:
                    if isinstance(value, pd.Timestamp):
                        display_value = value.strftime("%d/%m/%Y %H:%M:%S")
                    else:
                        display_value = str(value)
                item = QTableWidgetItem(display_value)

                # Bloqueia edição da coluna SEQ
                if self.df_filtered.columns[col] == 'SEQ':
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                # Cor de fundo conforme LOCALIZAÇÃO (opcional)
                if self.df_filtered.columns[col] == 'LOCALIZAÇÃO':
                    color_map = {
                        'Unidade': '#17a2b8',
                        'Escritório': '#ffc107',
                        'Sobreaviso': '#fd7e14',
                        'Folga': '#6c757d',
                        'Home': '#20c997',
                        'Online': '#8fbc8f'
                    }
                    cor = color_map.get(display_value, None)
                    if cor:
                        item.setBackground(QColor(cor))

                self.table_widget.setItem(row, col, item)

    def handle_header_click(self, logicalIndex):
        column_name = self.df_filtered.columns[logicalIndex]
        existing = next((item for item in self.sort_columns if item[0] == column_name), None)
        if existing:
            self.sort_columns.remove(existing)
            ascending = not existing[1]
        else:
            ascending = True

        self.sort_columns.insert(0, (column_name, ascending))
        order = Qt.AscendingOrder if ascending else Qt.DescendingOrder
        self.table_widget.horizontalHeader().setSortIndicator(logicalIndex, order)
        self.sort_table()

    def sort_table(self):
        if not self.sort_columns:
            return
        sort_by = [col for col, asc in self.sort_columns]
        ascending = [asc for col, asc in self.sort_columns]
        self.df_filtered.sort_values(
            by=sort_by, ascending=ascending, inplace=True,
            key=lambda col: col.str.lower() if col.dtype == object else col
        )
        self.df_filtered.reset_index(drop=True, inplace=True)
        self.populate_table()

    def enable_editing(self):
        self.table_widget.setEditTriggers(QTableWidget.AllEditTriggers)

    def delete_entry(self):
        selected_rows = self.table_widget.selectionModel().selectedRows()
        if selected_rows:
            selected_row = selected_rows[0].row()

            # Captura SEQ da linha que será removida, se existir
            if 'SEQ' in self.df_filtered.columns:
                seq_val = self.df_filtered.loc[self.df_filtered.index[selected_row], 'SEQ']
            else:
                seq_val = None

            self.table_widget.removeRow(selected_row)
            self.df_filtered.drop(self.df_filtered.index[selected_row], inplace=True)
            self.df_filtered.reset_index(drop=True, inplace=True)

            # Se SEQ for válido, guarda no set para exclusão definitiva
            if seq_val is not None and pd.notna(seq_val):
                self.deleted_seq.add(int(seq_val))

            QMessageBox.information(self, "Sucesso", "Entrada excluída com sucesso.")
        else:
            QMessageBox.warning(self, "Aviso", "Nenhuma linha selecionada para excluir.")

    def save_changes(self):
        # Atualiza self.df_filtered com o que está na tabela
        for row_index in range(self.table_widget.rowCount()):
            for col_index in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row_index, col_index)
                if item:
                    column_name = self.df_filtered.columns[col_index]
                    value = item.text()
                    if value == '':
                        if column_name in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
                            value = pd.NaT
                        else:
                            value = pd.NA
                    else:
                        if column_name in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
                            try:
                                value = pd.to_datetime(value, dayfirst=True, errors='raise')
                            except ValueError:
                                QMessageBox.warning(
                                    self,
                                    "Data/Hora Inválida",
                                    f"Formato de data/hora inválido para {column_name} na linha {row_index + 1}."
                                )
                                value = self.df_filtered.iloc[row_index, col_index]
                        elif column_name == 'SEQ':
                            try:
                                value = int(value)
                            except ValueError:
                                QMessageBox.warning(
                                    self, "Valor Inválido",
                                    f"Valor inválido para SEQ na linha {row_index + 1}."
                                )
                                value = self.df_filtered.iloc[row_index, col_index]
                        else:
                            value = str(value)
                    self.df_filtered.iloc[row_index, col_index] = value

        # >>> Remove do df_existing as linhas de SEQ que foram excluídas
        if 'SEQ' in self.df_existing.columns and len(self.deleted_seq) > 0:
            self.df_existing = self.df_existing[~self.df_existing['SEQ'].isin(self.deleted_seq)]
            # se quiser, pode limpar: self.deleted_seq.clear()

        # Agora, apenas atualizamos/adicionamos
        for idx in self.df_filtered.index:
            seq = self.df_filtered.loc[idx, 'SEQ']
            if pd.isna(seq):
                if not self.df_existing.empty:
                    max_seq = self.df_existing['SEQ'].max()
                else:
                    max_seq = 0
                if pd.isna(max_seq):
                    max_seq = 0
                else:
                    max_seq = int(max_seq)
                seq = max_seq + 1
                self.df_filtered.at[idx, 'SEQ'] = seq

            # Atualiza se já existir:
            if (self.df_existing['SEQ'] == seq).any():
                for col in self.df_filtered.columns:
                    if col != 'SEQ':
                        self.df_existing.loc[self.df_existing['SEQ'] == seq, col] = self.df_filtered.loc[idx, col]
            else:
                # Se for linha nova
                self.df_existing = pd.concat([self.df_existing, self.df_filtered.loc[[idx]]], ignore_index=True)

        for col in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
            self.df_existing[col] = pd.to_datetime(self.df_existing[col], errors='coerce')

        try:
            with pd.ExcelWriter(self.planilha_path, engine='openpyxl') as writer:
                self.df_existing.to_excel(writer, index=False)

            wb = load_workbook(self.planilha_path)
            ws = wb.active
            columns = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
            inicio_col = columns.get('DATA/HORA INICIO')
            fim_col = columns.get('DATA/HORA FIM')
            date_format = 'dd/mm/yyyy hh:mm:ss'

            if inicio_col:
                for cell in ws.iter_cols(min_col=inicio_col, max_col=inicio_col, min_row=2):
                    for c in cell:
                        c.number_format = date_format
            if fim_col:
                for cell in ws.iter_cols(min_col=fim_col, max_col=fim_col, min_row=2):
                    for c in cell:
                        c.number_format = date_format

            wb.save(self.planilha_path)
            QMessageBox.information(self, "Sucesso", f"Alterações salvas com sucesso em {self.planilha_path}!")
            self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)

            # Agora recarrega a planilha para sincronizar tudo
            self.reload_after_save()

        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Falha ao salvar as alterações: {e}")

    def reload_after_save(self):
        """
        Recarrega a planilha atualizada, reflete em df_existing, df_filtered e original_df.
        """
        try:
            updated_df = pd.read_excel(self.planilha_path)
            updated_df.columns = updated_df.columns.str.upper()

            # Converte colunas de data/hora
            for c in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
                if c in updated_df.columns:
                    updated_df[c] = pd.to_datetime(updated_df[c], format="%d/%m/%Y %H:%M:%S", errors='coerce')

            self.df_existing = updated_df.copy()
            self.original_df = updated_df.copy()
            self.df_filtered = updated_df.copy()

            # Mantém somente colunas necessárias
            self.df_filtered = self.df_filtered[self.labels]

            self.populate_table()

        except Exception as re_load_err:
            QMessageBox.warning(self, "Aviso", f"Falha ao recarregar planilha após salvar: {re_load_err}")

    def send_emails(self):
        """Abre EmailSelectionDialog p/ substring + período e envia e-mail."""
        tecnicos = list(self.df_filtered['TÉCNICO'].unique())
        dialog = EmailSelectionDialog(tecnicos, self.periodo_inicio, self.periodo_fim)
        if dialog.exec_() == QDialog.Accepted:
            typed_text = dialog.selected_tecnicos[0]
            periodo_inicio = dialog.periodo_inicio
            periodo_fim = dialog.periodo_fim

            try:
                df_existing = pd.read_excel(self.planilha_path)
                df_existing.columns = df_existing.columns.str.upper()
            except FileNotFoundError:
                QMessageBox.warning(self, "Erro", "A planilha selecionada não foi encontrada.")
                return

            if df_existing.empty:
                QMessageBox.information(self, "Aviso", "Não há dados na planilha para enviar.")
                return

            df_existing['DATA/HORA INICIO'] = pd.to_datetime(
                df_existing['DATA/HORA INICIO'], format="%d/%m/%Y %H:%M:%S", errors='coerce'
            )
            start_date = QDateTime(periodo_inicio, QTime(0, 0)).toPyDateTime()
            end_date = QDateTime(periodo_fim, QTime(23, 59, 59)).toPyDateTime()

            df_filtered = df_existing[
                (df_existing['DATA/HORA INICIO'] >= start_date) &
                (df_existing['DATA/HORA INICIO'] <= end_date) &
                (df_existing['TÉCNICO'].str.contains(typed_text, case=False, na=False))
            ]

            if df_filtered.empty:
                QMessageBox.information(self, "Aviso", "Não há registros para o técnico/período selecionados.")
                return

            try:
                locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
            except:
                locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')

            grouped = df_filtered.groupby('TÉCNICO')
            for tecnico, group in grouped:
                email = technician_emails.get(tecnico)
                if not email:
                    QMessageBox.warning(self, "Aviso", f"E-mail do técnico {tecnico} não encontrado.")
                    continue

                message = f"""
                <html>
                    <body>
                        <p>Prezado(a) <b>{tecnico}</b>,</p>
                        <p>Segue sua escala:</p>
                        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
                            <tr>
                                <th>DIA DA SEMANA</th>
                                <th>DATA</th>
                                <th>UNIDADE</th>
                                <th>LOCALIZAÇÃO</th>
                                <th>JUSTIFICATIVA</th>
                                <th>CARD</th>
                            </tr>
                """
                for idx, row in group.iterrows():
                    dia_semana = row.get('DIA DA SEMANA', '')
                    data_hora_inicio = row.get('DATA/HORA INICIO')
                    data_str = data_hora_inicio.strftime('%d/%m/%Y') if pd.notnull(data_hora_inicio) else ''
                    unidade = row.get('UNIDADE', '')
                    localizacao = row.get('LOCALIZAÇÃO', '')
                    justificativa = row.get('JUSTIFICATIVA', '')
                    card = row.get('CARD', '')

                    message += f"""
                            <tr>
                                <td>{dia_semana}</td>
                                <td>{data_str}</td>
                                <td>{unidade}</td>
                                <td>{localizacao}</td>
                                <td>{justificativa}</td>
                                <td>{card}</td>
                            </tr>
                    """
                message += """
                        </table>
                        <br>
                        <img src="cid:MinhaImagem" alt="Assinatura" />
                    </body>
                </html>
                """
                send_email(email, "Escala Semanal", message)

            reply = QMessageBox.question(
                self,
                'Enviar para Gestores',
                'Deseja enviar e-mails para os gestores das unidades?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                df_units = df_filtered.copy()
                df_units['DATA'] = df_units['DATA/HORA INICIO'].dt.date
                unit_grouped = df_units.groupby(['UNIDADE', 'DATA'])
                for (unidade, data_visita), group in unit_grouped:
                    gestor_email = unit_manager_emails.get(unidade)
                    if not gestor_email:
                        QMessageBox.warning(self, "Aviso", f"E-mail do gestor da unidade {unidade} não encontrado.")
                        continue

                    tecnicos_na_unidade = group['TÉCNICO'].unique()
                    tecnicos_lista = ', '.join(tecnicos_na_unidade)
                    data_inicio = group['DATA/HORA INICIO'].iloc[0]
                    horario_visita = data_inicio.strftime('%H:%M')

                    mensagem = f"""
                    <html>
                        <body>
                            <p>Prezado(a) Gestor(a),<br><br>
                            Informamos que o(s) técnico(s) <b>{tecnicos_lista}</b> estará(ão) presente(s) na unidade <b>{unidade}</b>
                            no dia {data_visita.strftime('%d/%m/%Y')} às {horario_visita}.<br>
                            <br>
                            Atenciosamente,<br>Sua Equipe
                            <br><br>
                            <img src="cid:MinhaImagem" alt="Assinatura" />
                        </body>
                    </html>
                    """
                    data_envio = datetime.datetime.combine(data_visita, datetime.time(7, 0))
                    try:
                        send_email(
                            gestor_email,
                            f"Visita de Técnico - {data_visita.strftime('%d/%m/%Y')}",
                            mensagem,
                            send_time=data_envio
                        )
                    except Exception as e:
                        QMessageBox.warning(self, "Erro", f"Falha ao agendar e-mail para {gestor_email}: {e}")

                QMessageBox.information(self, "Sucesso", "E-mails enviados aos técnicos e gestores das unidades.")
            else:
                QMessageBox.information(self, "Sucesso", "E-mails enviados apenas aos técnicos.")

    def get_primary_button_style(self):
        return """
            QPushButton {
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """

    def get_warning_button_style(self):
        return """
            QPushButton {
                background-color: #ffc107;
                color: #212529;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
        """

    def get_danger_button_style(self):
        return """
            QPushButton {
                background-color: #dc3545;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #bd2130;
            }
        """

# --------------------------------------------------
#        FORM PRINCIPAL PARA INCLUIR ESCALAS
# --------------------------------------------------
class ScheduleForm(QWidget):
    """
    Tela principal para Incluir Escala.
    Ao clicar em "Consultar Escala", iremos ignorar o período inicial
    e mostrar todos os dados da planilha em 'ConsultaEscalaDialog'.
    """
    def __init__(self, planilha_path, periodo_inicio, periodo_fim):
        super().__init__()
        self.setWindowIcon(QIcon('JT.ico'))
        self.planilha_path = planilha_path
        self.periodo_inicio = periodo_inicio
        self.periodo_fim = periodo_fim
        self.editing_row = None
        self.enter_pressed_once = False
        self.enter_timer = QTimer(self)
        self.enter_timer.setInterval(500)
        self.enter_timer.timeout.connect(self.reset_enter)
        self.is_editing_entry = False
        self.tecnicos_12x36_dias = {}

        # Carrega escalas_tecnicos.json
        try:
            with open('escalas_tecnicos.json', 'r', encoding='utf-8') as f:
                self.technician_schedules = json.load(f)
        except FileNotFoundError:
            QMessageBox.critical(self, "Erro", "Arquivo 'escalas_tecnicos.json' não encontrado.")
            sys.exit()
        except json.JSONDecodeError as e:
            QMessageBox.critical(self, "Erro", f"Erro ao decodificar o arquivo JSON: {e}")
            sys.exit()

        self.init_ui()

    def init_ui(self):
        """
        Layout completo do ScheduleForm,
        incluindo a tabela, combos de localização/unidade/tecnico,
        e localizacao_options com "Online".
        """
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.setSpacing(0)

        self.periodo_label = ClickableLabel(
            f"Período Selecionado: {self.periodo_inicio.toString('dd/MM/yyyy')} a {self.periodo_fim.toString('dd/MM/yyyy')}"
        )
        self.periodo_label.setAlignment(Qt.AlignCenter)
        self.periodo_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                margin: 10px;
                color: #343a40;
            }
            QLabel:hover {
                color: #007bff;
                text-decoration: underline;
                cursor: pointer;
            }
        """)
        self.periodo_label.clicked.connect(self.change_period)
        main_layout.addWidget(self.periodo_label)

        self.labels = [
            'DIA DA SEMANA', 'LOCALIZAÇÃO', 'UNIDADE', 'TÉCNICO', 'ESCALA', 'TURNO',
            'DATA/HORA INICIO', 'DATA/HORA FIM', 'JUSTIFICATIVA', 'CARD'
        ]

        localizacao_options = ["", "Folga", "Férias", "Sobreaviso", "Unidade", "Escritório", "Home", "Online"]

        # Exemplo de unidades. Pode adaptar conforme necessidade.
        unidades = [
    "CRST Freguesia do Ó", 
    "CRST Lapa", 
    "CRST Leste", 
    "CRST Mooca", 
    "CRST Santo Amaro", 
    "CRST Sé", 
    "H Cantareira", 
    "HD Brasilandia", 
    "HM Alipio", 
    "HM Benedicto",

    "HM Brigadeiro", 
    "HM Cachoeirinha", 
    "HM Campo Limpo", 
    "HM Capela Do Socorro", 
    "HM Hungria", 
    "HM Ignacio", 
    "HM Mario Degni", 
    "HM Saboya", 
    "HM Sorocabana", 
    "HM Tatuape",

    "HM Tide", 
    "HM Waldomiro", 
    "HM Zaio", 
    "PA Sao Mateus", 
    "PSM Balneario Sao Jose", 
    "UPA Dona Maria Antonieta", 
    "UPA Elisa Maria", 
    "UPA Lapa", 
    "UPA Parelheiros", 
    "UPA Pedreira",

    "CAPS AD II Cachoeirinha", 
    "CAPS AD II Cangaiba", 
    "CAPS AD II Cidade Ademar", 
    "CAPS AD II Ermelino Matarazzo", 
    "CAPS AD II Guaianases", 
    "CAPS AD II Jabaquara", 
    "CAPS AD II Jardim Nelia", 
    "CAPS AD II Mooca", 
    "CAPS AD II Pinheiros", 
    "CAPS AD II Sacoma",

    "CAPS AD II Santo Amaro", 
    "CAPS AD II Sapopemba", 
    "CAPS AD II Vila Madalena Prosam", 
    "CAPS AD II Vila Mariana", 
    "CAPS AD III Armenia", 
    "CAPS AD III Boracea", 
    "CAPS AD III Butanta", 
    "CAPS AD III Campo Limpo", 
    "CAPS AD III Capela Do Socorro", 
    "CAPS AD III Centro",

    "CAPS AD III Complexo Prates", 
    "CAPS AD III Freguesia Do O Brasilandia", 
    "CAPS AD III Grajau", 
    "CAPS AD III Heliopolis", 
    "CAPS AD III Itaquera", 
    "CAPS AD III Jardim Angela", 
    "CAPS AD III Jardim Sao Luiz", 
    "CAPS AD III Leopoldina", 
    "CAPS AD III Mandaqui", 
    "CAPS AD III Paraisopolis",

    "CAPS AD III Penha", 
    "CAPS AD III Pirituba Casa Azul", 
    "CAPS AD III Santana", 
    "CAPS AD III Sao Mateus Liberdade De Escolha", 
    "CAPS AD III Sao Miguel", 
    "CAPS AD IV Redencao", 
    "CAPS Adulto II Aricanduva Formosa", 
    "CAPS Adulto II Brasilandia", 
    "CAPS Adulto II Butanta", 
    "CAPS Adulto II Casa Verde",

    "CAPS Adulto II Cidade Ademar", 
    "CAPS Adulto II Cidade Tiradentes", 
    "CAPS Adulto II Ermelino Matarazzo", 
    "CAPS Adulto II Guaianases Artur Bispo Do Rosario", 
    "CAPS Adulto II Itaim Bibi", 
    "CAPS Adulto II Itaim Paulista", 
    "CAPS Adulto II Itaquera", 
    "CAPS Adulto II Jabaquara", 
    "CAPS Adulto II Jacana Dr Leonidio Galvao Dos Santos", 
    "CAPS Adulto II Jardim Lidia",

    "CAPS Adulto II Perdizes Manoel Munhoz", 
    "CAPS Adulto II Perus", 
    "CAPS Adulto II Sao Miguel", 
    "CAPS Adulto II Vila Monumento", 
    "CAPS Adulto II Vila Prudente", 
    "CAPS Adulto III Capela Do Socorro", 
    "CAPS Adulto III Freguesia Do O Brasilandia", 
    "CAPS Adulto III Grajau", 
    "CAPS Adulto III Itaim Bibi", 
    "CAPS Adulto III Jardim Sao Luiz",

    "CAPS Adulto III Lapa", 
    "CAPS Adulto III Largo 13", 
    "CAPS Adulto III M Boi Mirim", 
    "CAPS Adulto III Mandaqui", 
    "CAPS Adulto III Mooca", 
    "CAPS Adulto III Paraisopolis", 
    "CAPS Adulto III Parelheiros", 
    "CAPS Adulto III Perdizes", 
    "CAPS Adulto III Pirituba Jaragua", 
    "CAPS Adulto III Sao Mateus",

    "CAPS Adulto III Sapopemba", 
    "CAPS Adulto III Se", 
    "CAPS Adulto III Vila Matilde", 
    "CAPS IJ II Butanta", 
    "CAPS IJ II Campo Limpo", 
    "CAPS IJ II Capela Do Socorro Piracao", 
    "CAPS IJ II Casa Verde Nise Da Silveira", 
    "CAPS IJ II Cidade Ademar", 
    "CAPS IJ II Cidade Lider", 
    "CAPS IJ II Cidade Tiradentes",

    "CAPS IJ II Ermelino Matarazzo", 
    "CAPS IJ II Freguesia Do O Brasilandia", 
    "CAPS IJ II Guaianases Coloridamente", 
    "CAPS IJ II Ipiranga", 
    "CAPS IJ II Itaim Paulista", 
    "CAPS IJ II Itaquera", 
    "CAPS IJ II Jabaquara Casinha", 
    "CAPS IJ II Lapa", 
    "CAPS IJ II M Boi Mirim", 
    "CAPS IJ II Mooca",

    "CAPS IJ II Parelheiros Aquarela", 
    "CAPS IJ II Perus", 
    "CAPS IJ II Pirituba Jaragua", 
    "CAPS IJ II Santo Amaro", 
    "CAPS IJ II Sao Mateus", 
    "CAPS IJ II Sapopemba", 
    "CAPS IJ II Vila Maria Vila Guilherme", 
    "CAPS IJ II Vila Mariana Quixote", 
    "CAPS IJ II Vila Prudente", 
    "CAPS IJ III Aricanduva",

    "CAPS IJ III Cidade Dutra", 
    "CAPS IJ III Heliopolis", 
    "CAPS IJ III Jardim Sao Luiz", 
    "CAPS IJ III Penha", 
    "CAPS IJ III Pirituba", 
    "CAPS IJ III Santana", 
    "CAPS IJ III Sao Miguel", 
    "CAPS IJ III Se Amorzeira", 
    "CIES - Luz Campos Elíseos", 
    "PSM Álvaro Dino",

    "CNR Redenção", 
    "UPA 21 de Junho", 
    "UPA 26 de Agosto", 
    "SCP AD Boracea", 
    "SCP AD Pirituba", 
    "UPA Campo Limpo", 
    "UPA Carrao", 
    "UPA City Jaragua", 
    "UPA Ermelino Matarazzo", 
    "UPA Jabaquara",

    "UPA Jacana", 
    "UPA Jardim Angela", 
    "UPA Julio Tupy", 
    "UPA Mooca", 
    "UPA Parque Doroteia", 
    "UPA Peri", 
    "UPA Perus", 
    "UPA Pirituba", 
    "UPA Rio Pequeno", 
    "UPA Santo Amaro",

    "UPA Tatuape", 
    "UPA Tiradentes", 
    "UPA Tito Lopes", 
    "UPA Vera Cruz", 
    "UPA Vergueiro", 
    "UPA Vila Mariana", 
    "UPA Vila Santa Catarina", 
    "Ouro Verde", 
    "UPA Anchieta", 
    "UPA Campo Grande",

    "UPA Carlos Lourenço", 
    "UPA Sao Jose",
    "UPA Barra Funda",
    "UPA Augusto Gomes de Matos"
]

        tecnicos = list(self.technician_schedules.keys())
        turnos = ["Diurno", "Noturno"]

        self.dia_semana = QLabel("-")
        self.dia_semana.setAlignment(Qt.AlignCenter)
        self.dia_semana.setStyleSheet("""
            QLabel {
                border: 1px solid #ced4da;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
                border-radius: 4px;
            }
        """)

        self.combo_box_localizacao = QComboBox()
        self.combo_box_localizacao.addItems(localizacao_options)
        self.combo_box_localizacao.setCurrentIndex(0)
        self.combo_box_localizacao.currentIndexChanged.connect(self.handle_localizacao_change)
        self.combo_box_localizacao.setStyleSheet("""
            QComboBox {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
            QComboBox QAbstractItemView {
                background-color: #fff;
                selection-background-color: #007bff;
                selection-color: #fff;
            }
        """)

        self.combo_box_unidade = FilteredComboBox(unidades, parent=self)
        self.combo_box_tecnico = FilteredComboBox(tecnicos, parent=self)
        self.combo_box_turno = FilteredComboBox(turnos, parent=self)

        self.combo_box_tecnico.lineEdit().editingFinished.connect(self.update_fields_based_on_tecnico)
        self.combo_box_unidade.lineEdit().editingFinished.connect(self.update_fields_based_on_tecnico)

        self.escala_label = QLabel("-")
        self.escala_label.setAlignment(Qt.AlignCenter)
        self.escala_label.setStyleSheet("""
            QLabel {
                border: 1px solid #ced4da;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
                border-radius: 4px;
            }
        """)

        self.date_time_edit_inicio = QDateTimeEdit()
        self.date_time_edit_fim = QDateTimeEdit()

        self.date_time_edit_inicio.setCalendarPopup(True)
        self.date_time_edit_inicio.setDisplayFormat("dd/MM/yyyy HH:mm:ss")
        self.date_time_edit_inicio.setStyleSheet("""
            QDateTimeEdit {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
            QDateTimeEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #ced4da;
            }
        """)
        self.date_time_edit_inicio.setMinimumDateTime(QDateTime(self.periodo_inicio, QTime(0, 0, 0)))
        self.date_time_edit_inicio.setMaximumDateTime(QDateTime(self.periodo_fim, QTime(23, 59, 59)))
        self.date_time_edit_inicio.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))
        self.date_time_edit_inicio.dateTimeChanged.connect(self.update_dia_semana_from_datetime)

        self.date_time_edit_fim.setCalendarPopup(True)
        self.date_time_edit_fim.setDisplayFormat("dd/MM/yyyy HH:mm:ss")
        self.date_time_edit_fim.setStyleSheet("""
            QDateTimeEdit {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
            QDateTimeEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #ced4da;
            }
        """)
        self.date_time_edit_fim.setMinimumDate(self.periodo_inicio)
        self.date_time_edit_fim.setMaximumDate(self.periodo_fim.addDays(1))
        self.date_time_edit_fim.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))

        self.justificativa = QLineEdit()
        self.justificativa.setStyleSheet("""
            QLineEdit {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
        """)

        self.card_input = QLineEdit()
        self.card_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
                background-color: #fff;
            }
        """)

        self.input_widgets = [
            self.dia_semana,
            self.combo_box_localizacao,
            self.combo_box_unidade,
            self.combo_box_tecnico,
            self.escala_label,
            self.combo_box_turno,
            self.date_time_edit_inicio,
            self.date_time_edit_fim,
            self.justificativa,
            self.card_input
        ]

        self.form_table = QTableWidget()
        self.form_table.setRowCount(1)
        self.form_table.setColumnCount(len(self.labels))
        self.form_table.setHorizontalHeaderLabels(self.labels)
        self.form_table.verticalHeader().setVisible(False)
        self.form_table.horizontalHeader().setStretchLastSection(True)
        self.form_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.form_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.form_table.setSelectionMode(QTableWidget.NoSelection)
        self.form_table.setFocusPolicy(Qt.NoFocus)
        self.form_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.form_table.setMaximumHeight(100)
        self.form_table.setStyleSheet("""
            QTableWidget {
                border: none;
            }
            QHeaderView::section {
                background-color: qlineargradient(
                    spread:pad, x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6959CD, stop:1 #FFFFFF
                );
                color: black;
                font-weight: bold;
                font-size: 15px;
                font-family: 'Segoe UI', sans-serif;
                height: 35px;
            }
        """)

        for col, widget in enumerate(self.input_widgets):
            self.form_table.setCellWidget(0, col, widget)

        main_layout.addWidget(self.form_table)

        self.form_table.horizontalHeader().setSectionsClickable(True)
        self.form_table.horizontalHeader().setSortIndicatorShown(True)
        self.form_table.horizontalHeader().sectionClicked.connect(self.handle_header_click)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(len(self.labels))
        self.table_widget.horizontalHeader().setVisible(False)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                border: none;
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
            }
            QTableWidget::item {
                border-bottom: 1px solid #dee2e6;
            }
            QTableWidget::item:selected {
                background-color: #007bff;
                color: #fff;
            }
        """)
        self.table_widget.setShowGrid(False)
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table_widget)

        self.sort_states = {2: False, 3: False, 6: False}
        self.original_data = []

        buttons_layout = QHBoxLayout()
        left_buttons_layout = QHBoxLayout()
        left_buttons_layout.setSpacing(10)

        self.consult_button = QPushButton(" Consultar Escala")
        self.consult_button.setFixedSize(180, 40)
        self.consult_button.setIcon(QIcon("icons/consult.png"))
        self.consult_button.setStyleSheet(self.get_primary_button_style())
        self.consult_button.clicked.connect(self.consultar_escala)
        left_buttons_layout.addWidget(self.consult_button)

        self.incluir_escala_button = QPushButton(" Incluir Escala Pré-Preenchida")
        self.incluir_escala_button.setFixedSize(220, 40)
        self.incluir_escala_button.setIcon(QIcon("icons/add_schedule.png"))
        self.incluir_escala_button.setStyleSheet(self.get_success_button_style())
        self.incluir_escala_button.clicked.connect(self.incluir_escala_semanal)
        left_buttons_layout.addWidget(self.incluir_escala_button)

        self.send_email_button = QPushButton(" Enviar Escala")
        self.send_email_button.setFixedSize(180, 40)
        self.send_email_button.setIcon(QIcon("icons/email.png"))
        self.send_email_button.setStyleSheet(self.get_primary_button_style())
        self.send_email_button.clicked.connect(self.send_emails)
        left_buttons_layout.addWidget(self.send_email_button)

        buttons_layout.addLayout(left_buttons_layout)
        buttons_layout.addStretch()

        right_buttons_layout = QHBoxLayout()
        right_buttons_layout.setSpacing(10)

        self.add_button = QPushButton(" Adicionar")
        self.add_button.setFixedSize(140, 40)
        self.add_button.setIcon(QIcon("icons/add.png"))
        self.add_button.setStyleSheet(self.get_success_button_style())
        self.add_button.clicked.connect(self.add_entry)
        right_buttons_layout.addWidget(self.add_button)

        self.edit_button = QPushButton(" Editar")
        self.edit_button.setFixedSize(140, 40)
        self.edit_button.setIcon(QIcon("icons/edit.png"))
        self.edit_button.setStyleSheet(self.get_warning_button_style())
        self.edit_button.clicked.connect(self.edit_entry)
        right_buttons_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton(" Excluir")
        self.delete_button.setFixedSize(140, 40)
        self.delete_button.setIcon(QIcon("icons/delete.png"))
        self.delete_button.setStyleSheet(self.get_danger_button_style())
        self.delete_button.clicked.connect(self.delete_entry)
        right_buttons_layout.addWidget(self.delete_button)

        self.finalize_button = QPushButton(" Finalizar Escala")
        self.finalize_button.setFixedSize(180, 40)
        self.finalize_button.setIcon(QIcon("icons/save.png"))
        self.finalize_button.setStyleSheet(self.get_primary_button_style())
        self.finalize_button.clicked.connect(self.finalize_schedule)
        right_buttons_layout.addWidget(self.finalize_button)

        buttons_layout.addLayout(right_buttons_layout)
        main_layout.addLayout(buttons_layout)

        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
            }
            QPushButton {
                font-size: 14px;
                font-family: 'Segoe UI', sans-serif;
            }
        """)
        self.setLayout(main_layout)
        self.setWindowTitle('GIRA TURNOS')
        self.resize(1200, 600)
        self.show()

    def get_primary_button_style(self):
        return """
            QPushButton {
                background-color: #6959CD;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """

    def get_success_button_style(self):
        return """
            QPushButton {
                background-color: #6959CD;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1e7e34;
            }
        """

    def get_warning_button_style(self):
        return """
            QPushButton {
                background-color: #6959CD;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
        """

    def get_danger_button_style(self):
        return """
            QPushButton {
                background-color: #6959CD;
                color: white;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #bd2130;
            }
        """

    def change_period(self):
        """Permite mudar o período durante a inclusão de escala."""
        periodo_dialog = PeriodoConsultaDialog()
        if periodo_dialog.exec_() == QDialog.Accepted:
            self.periodo_inicio = periodo_dialog.periodo_inicio
            self.periodo_fim = periodo_dialog.periodo_fim
            self.periodo_label.setText(
                f"Período Selecionado: {self.periodo_inicio.toString('dd/MM/yyyy')} a {self.periodo_fim.toString('dd/MM/yyyy')}"
            )
            self.date_time_edit_inicio.setMinimumDate(self.periodo_inicio)
            self.date_time_edit_inicio.setMaximumDate(self.periodo_fim)
            self.date_time_edit_fim.setMinimumDate(self.periodo_inicio)
            self.date_time_edit_fim.setMaximumDate(self.periodo_fim.addDays(1))

    def reset_enter(self):
        self.enter_pressed_once = False
        self.enter_timer.stop()

    def handle_enter_press(self):
        if self.enter_pressed_once:
            self.enter_pressed_once = False
            self.enter_timer.stop()
            self.add_entry()
        else:
            self.enter_pressed_once = True
            self.enter_timer.start()

    def should_work(self, tecnico, date):
        """Verifica se técnico deve trabalhar numa data (5x2 ou 12x36)."""
        tecnico_info = self.technician_schedules.get(tecnico, {})
        if not tecnico_info:
            return False
        escala = tecnico_info.get('escala')
        if escala == '5X2':
            dia_semana = date.dayOfWeek()  # 1=Seg ... 7=Dom
            return dia_semana in tecnico_info.get('dias_trabalho', [])
        elif escala == '12X36':
            dias_trabalho = self.tecnicos_12x36_dias.get(tecnico)
            if dias_trabalho:
                dia_mes = date.day()
                if dias_trabalho == 'pares':
                    return (dia_mes % 2 == 0)
                elif dias_trabalho == 'impares':
                    return (dia_mes % 2 != 0)
        return False

    def does_on_call(self, tecnico):
        """Verifica se técnico faz sobreaviso."""
        tecnico_info = self.technician_schedules.get(tecnico)
        if not tecnico_info:
            return False
        sobreaviso_info = tecnico_info.get('sobreaviso', {})
        return sobreaviso_info.get('faz_sobreaviso', False)

    def incluir_escala_semanal(self):
        """Inclui escala pré-preenchida para o período selecionado."""
        tecnicos = list(self.technician_schedules.keys())
        dialog = TechnicianSelectionDialog(tecnicos, self.technician_schedules)
        if dialog.exec_() == QDialog.Accepted:
            selected_tecnicos = dialog.selected_tecnicos
            self.tecnicos_12x36_dias = {}
            for tecnico, dias_trabalho in selected_tecnicos.items():
                if self.technician_schedules[tecnico]['escala'] == '12X36':
                    self.tecnicos_12x36_dias[tecnico] = dias_trabalho
            for tecnico in selected_tecnicos:
                current_date = self.periodo_inicio
                while current_date <= self.periodo_fim:
                    # Verifica apenas se precisa inserir (12x36 ou 5x2),
                    # mas não impede se quiser criar manualmente no fim de semana
                    if self.should_work(tecnico, current_date):
                        self.combo_box_tecnico.setCurrentText(tecnico)
                        self.date_time_edit_inicio.setDate(current_date)
                        self.update_fields_based_on_tecnico()
                        self.add_entry()
                    current_date = current_date.addDays(1)

    def add_entry(self):
        """Adiciona uma entrada na tabela."""
        dia_semana = self.dia_semana.text()
        localizacao = self.combo_box_localizacao.currentText()
        unidade = self.combo_box_unidade.currentText()
        tecnico = self.combo_box_tecnico.currentText()
        escala = self.escala_label.text()
        turno = self.combo_box_turno.currentText()
        data_hora_inicio = self.date_time_edit_inicio.dateTime()
        data_hora_fim = self.date_time_edit_fim.dateTime()
        justificativa = self.justificativa.text()
        card_value = self.card_input.text()

        # -------------------------------
        # *** AQUI FOI REMOVIDA A VALIDAÇÃO
        # QUE IMPEDIA TÉCNICO 5X2 NO FIM DE SEMANA
        # -------------------------------

        # Verificação se é sobreaviso e o técnico não faz sobreaviso:
        if localizacao == 'Sobreaviso' and not self.does_on_call(tecnico):
            QMessageBox.warning(
                self, "Erro",
                f"{tecnico} não está configurado para Sobreaviso."
            )
            return

        data_hora_inicio_dt = data_hora_inicio.toPyDateTime()
        data_hora_fim_dt = data_hora_fim.toPyDateTime()
        data_hora_inicio_str = data_hora_inicio_dt.strftime("%d/%m/%Y %H:%M:%S")
        data_hora_fim_str = data_hora_fim_dt.strftime("%d/%m/%Y %H:%M:%S")

        fields = [
            dia_semana,
            localizacao,
            unidade,
            tecnico,
            escala,
            turno,
            data_hora_inicio_str,
            data_hora_fim_str,
            justificativa,
            card_value
        ]

        if self.editing_row is not None:
            for column, value in enumerate(fields):
                self.table_widget.setItem(self.editing_row, column, QTableWidgetItem(value))
            self.original_data[self.editing_row] = fields
            self.editing_row = None
            self.add_button.setText(" Adicionar")
            self.add_button.setIcon(QIcon("icons/add.png"))
        else:
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            for column, value in enumerate(fields):
                self.table_widget.setItem(row_position, column, QTableWidgetItem(value))
            self.original_data.append(fields)

            # Exemplo de inserir uma folga automática após sobreaviso de domingo
            if localizacao == 'Sobreaviso':
                day_of_week = data_hora_inicio.date().dayOfWeek()
                if day_of_week == 7:  # Domingo
                    tecnico_info = self.technician_schedules.get(tecnico)
                    if tecnico_info:
                        folga_date = data_hora_inicio.addDays(1)
                        horario_inicio = tecnico_info.get('horario_inicio', '00:00')
                        horario_fim = tecnico_info.get('horario_fim', '00:00')
                        folga_inicio = QDateTime(folga_date.date(), QTime.fromString(horario_inicio, "HH:mm"))
                        folga_fim = QDateTime(folga_date.date(), QTime.fromString(horario_fim, "HH:mm"))

                        folga_dia_semana = self.get_dia_semana_text(folga_inicio)
                        folga_fields = [
                            folga_dia_semana,
                            "Folga",
                            "",
                            tecnico,
                            escala,
                            turno,
                            folga_inicio.toString("dd/MM/yyyy HH:mm:ss"),
                            folga_fim.toString("dd/MM/yyyy HH:mm:ss"),
                            "Folga após sobreaviso",
                            ""
                        ]
                        row_position = self.table_widget.rowCount()
                        self.table_widget.insertRow(row_position)
                        for column, value2 in enumerate(folga_fields):
                            self.table_widget.setItem(row_position, column, QTableWidgetItem(value2))
                        self.original_data.append(folga_fields)

        self.clear_fields()

    def clear_fields(self):
        self.combo_box_localizacao.setCurrentIndex(0)
        self.combo_box_unidade.setCurrentIndex(0)
        self.combo_box_turno.setCurrentIndex(0)
        self.date_time_edit_inicio.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))
        self.date_time_edit_fim.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))
        self.justificativa.clear()
        self.card_input.clear()
        self.dia_semana.setText("-")
        self.escala_label.setText("-")
        self.add_button.setText(" Adicionar")
        self.add_button.setIcon(QIcon("icons/add.png"))
        self.editing_row = None

    def edit_entry(self):
        selected_rows = self.table_widget.selectionModel().selectedRows()
        if selected_rows:
            selected_row = selected_rows[0].row()
            self.editing_row = selected_row
            self.is_editing_entry = True

            try:
                self.combo_box_tecnico.lineEdit().editingFinished.disconnect(self.update_fields_based_on_tecnico)
                self.combo_box_unidade.lineEdit().editingFinished.disconnect(self.update_fields_based_on_tecnico)
            except TypeError:
                pass

            dia_semana = self.table_widget.item(selected_row, 0).text()
            localizacao = self.table_widget.item(selected_row, 1).text()
            unidade = self.table_widget.item(selected_row, 2).text()
            tecnico = self.table_widget.item(selected_row, 3).text()
            escala = self.table_widget.item(selected_row, 4).text()
            turno = self.table_widget.item(selected_row, 5).text()
            data_hora_inicio = self.table_widget.item(selected_row, 6).text()
            data_hora_fim = self.table_widget.item(selected_row, 7).text()
            justificativa = self.table_widget.item(selected_row, 8).text()
            card_value = self.table_widget.item(selected_row, 9).text()

            self.dia_semana.setText(dia_semana)
            index_localizacao = self.combo_box_localizacao.findText(localizacao)
            if index_localizacao >= 0:
                self.combo_box_localizacao.setCurrentIndex(index_localizacao)
            self.combo_box_unidade.setCurrentText(unidade)
            self.combo_box_tecnico.setCurrentText(tecnico)
            self.escala_label.setText(escala)
            self.combo_box_turno.setCurrentText(turno)

            dt_inicio = QDateTime.fromString(data_hora_inicio, "dd/MM/yyyy HH:mm:ss")
            dt_fim = QDateTime.fromString(data_hora_fim, "dd/MM/yyyy HH:mm:ss")
            if dt_inicio.isValid():
                self.date_time_edit_inicio.setDateTime(dt_inicio)
            else:
                self.date_time_edit_inicio.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))
            if dt_fim.isValid():
                self.date_time_edit_fim.setDateTime(dt_fim)
            else:
                self.date_time_edit_fim.setDateTime(QDateTime(self.periodo_inicio, QTime.currentTime()))

            self.justificativa.setText(justificativa)
            self.card_input.setText(card_value)

            self.add_button.setText(" Gravar")
            self.add_button.setIcon(QIcon("icons/save.png"))

            self.combo_box_tecnico.lineEdit().editingFinished.connect(self.update_fields_based_on_tecnico)
            self.combo_box_unidade.lineEdit().editingFinished.connect(self.update_fields_based_on_tecnico)
            self.is_editing_entry = False
        else:
            QMessageBox.warning(self, "Aviso", "Nenhuma linha selecionada para editar.")

    def delete_entry(self):
        selected_rows = self.table_widget.selectionModel().selectedRows()
        if selected_rows:
            selected_row = selected_rows[0].row()
            self.table_widget.removeRow(selected_row)
            del self.original_data[selected_row]
            self.clear_fields()
        else:
            QMessageBox.warning(self, "Aviso", "Nenhuma linha selecionada para excluir.")

    def handle_header_click(self, logicalIndex):
        if logicalIndex in self.sort_states:
            self.sort_states[logicalIndex] = not self.sort_states[logicalIndex]
            order = Qt.AscendingOrder if self.sort_states[logicalIndex] else Qt.DescendingOrder
            self.form_table.horizontalHeader().setSortIndicator(logicalIndex, order)
            self.sort_table(logicalIndex, order)

    def sort_table(self, column, order):
        data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for col in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, col)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        reverse = (order == Qt.DescendingOrder)
        if column == 6:
            data.sort(key=lambda x: QDateTime.fromString(x[column], "dd/MM/yyyy HH:mm:ss").toPyDateTime(), reverse=reverse)
        else:
            data.sort(key=lambda x: x[column], reverse=reverse)

        self.table_widget.setRowCount(0)
        for row_data in data:
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            for column_index, value in enumerate(row_data):
                self.table_widget.setItem(row_position, column_index, QTableWidgetItem(value))
        self.original_data = data

    def finalize_schedule(self):
        if not self.original_data:
            QMessageBox.information(self, "Aviso", "Não há entradas para salvar.")
            return

        try:
            df_existing = pd.read_excel(self.planilha_path)
            df_existing.columns = df_existing.columns.str.upper()
        except FileNotFoundError:
            df_existing = pd.DataFrame()

        if not df_existing.empty and 'SEQ' in df_existing.columns:
            try:
                df_existing['SEQ'] = df_existing['SEQ'].astype(int)
            except ValueError:
                df_existing['SEQ'] = pd.to_numeric(df_existing['SEQ'], errors='coerce').fillna(0).astype(int)
        elif not df_existing.empty and 'SEQ' not in df_existing.columns:
            df_existing.insert(0, 'SEQ', range(1, len(df_existing) + 1))

        df_new = pd.DataFrame(self.original_data, columns=self.labels)

        if not df_existing.empty and 'SEQ' in df_existing.columns:
            max_seq = df_existing['SEQ'].max()
            if pd.isna(max_seq):
                max_seq = 0
            else:
                max_seq = int(max_seq)
        else:
            max_seq = 0

        df_new.insert(0, 'SEQ', range(max_seq + 1, max_seq + len(df_new) + 1))

        for col in ['DATA/HORA INICIO', 'DATA/HORA FIM']:
            df_new[col] = pd.to_datetime(df_new[col], format="%d/%m/%Y %H:%M:%S", errors='coerce')

        df_final = pd.concat([df_existing, df_new], ignore_index=True)
        df_final.to_excel(self.planilha_path, index=False, engine='openpyxl')

        wb = load_workbook(self.planilha_path)
        ws = wb.active
        columns = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        inicio_col = columns.get('DATA/HORA INICIO')
        fim_col = columns.get('DATA/HORA FIM')
        date_format = 'dd/mm/yyyy hh:mm:ss'

        if inicio_col:
            for cell in ws.iter_cols(min_col=inicio_col, max_col=inicio_col, min_row=2):
                for c in cell:
                    c.number_format = date_format

        if fim_col:
            for cell in ws.iter_cols(min_col=fim_col, max_col=fim_col, min_row=2):
                for c in cell:
                    c.number_format = date_format

        wb.save(self.planilha_path)

        self.table_widget.setRowCount(0)
        self.original_data.clear()
        QMessageBox.information(self, "Sucesso", "Escala finalizada e salva com sucesso!")

    def consultar_escala(self, periodo_inicio=None, periodo_fim=None):
        """
        Ignora o período e mostra TODOS os dados da planilha em ConsultaEscalaDialog.
        """
        try:
            df_existing = pd.read_excel(self.planilha_path)
            df_existing.columns = df_existing.columns.str.upper()
        except FileNotFoundError:
            QMessageBox.warning(self, "Erro", "A planilha selecionada não foi encontrada.")
            return

        if df_existing.empty:
            QMessageBox.information(self, "Aviso", "Não há dados na planilha para consultar.")
            return

        df_existing['DATA/HORA INICIO'] = pd.to_datetime(
            df_existing['DATA/HORA INICIO'], format="%d/%m/%Y %H:%M:%S", errors='coerce'
        )

        # Carrega tudo (sem filtrar pelo período)
        df_filtered = df_existing.copy()

        if df_filtered.empty:
            QMessageBox.information(self, "Aviso", "Não há registros na planilha.")
            return

        # Abre a nova tela de consulta com TUDO
        self.consulta_dialog = ConsultaEscalaDialog(
            df_filtered,
            self.planilha_path,
            df_existing,
            self.periodo_inicio,
            self.periodo_fim,
            ['SEQ'] + self.labels
        )
        self.consulta_dialog.show()

    def update_fields_based_on_tecnico(self):
        if self.is_editing_entry:
            return

        tecnico_nome = self.combo_box_tecnico.currentText()
        tecnico_info = self.technician_schedules.get(tecnico_nome)
        if not tecnico_info:
            self.escala_label.setText('-')
            return

        self.escala_label.setText(tecnico_info.get('escala', '-'))
        selected_localizacao = self.combo_box_localizacao.currentText()
        unidade_preenchida = bool(self.combo_box_unidade.currentText())

        selected_datetime = self.date_time_edit_inicio.dateTime()
        selected_date = selected_datetime.date()

        # Se estiver em sobreaviso
        if selected_localizacao == 'Sobreaviso' and self.does_on_call(tecnico_nome):
            sobreaviso_info = tecnico_info.get('sobreaviso', {})
            if unidade_preenchida:
                horario_inicio = sobreaviso_info.get('horario_com_unidade', {}).get('inicio', tecnico_info.get('horario_inicio', '00:00'))
                horario_fim = sobreaviso_info.get('horario_com_unidade', {}).get('fim', tecnico_info.get('horario_fim', '00:00'))
            else:
                horario_inicio = sobreaviso_info.get('horario_sem_unidade', {}).get('inicio', tecnico_info.get('horario_inicio', '00:00'))
                horario_fim = sobreaviso_info.get('horario_sem_unidade', {}).get('fim', tecnico_info.get('horario_fim', '00:00'))
        else:
            horario_inicio = tecnico_info.get('horario_inicio', '00:00')
            horario_fim = tecnico_info.get('horario_fim', '00:00')

        datetime_inicio_str = f"{selected_date.toString('dd/MM/yyyy')} {horario_inicio}"
        datetime_fim_str = f"{selected_date.toString('dd/MM/yyyy')} {horario_fim}"
        datetime_inicio = QDateTime.fromString(datetime_inicio_str, "dd/MM/yyyy HH:mm")
        datetime_fim = QDateTime.fromString(datetime_fim_str, "dd/MM/yyyy HH:mm")

        if datetime_fim <= datetime_inicio:
            datetime_fim = datetime_fim.addDays(1)

        self.date_time_edit_inicio.setDateTime(datetime_inicio)
        self.date_time_edit_fim.setDateTime(datetime_fim)

        self.update_dia_semana(datetime_inicio)
        hour_inicio = datetime_inicio.time().hour()
        if 6 <= hour_inicio < 18:
            self.combo_box_turno.setCurrentText("Diurno")
        else:
            self.combo_box_turno.setCurrentText("Noturno")

    def update_dia_semana_from_datetime(self):
        selected_datetime = self.date_time_edit_inicio.dateTime()
        self.update_dia_semana(selected_datetime)

    def get_dia_semana_text(self, selected_datetime):
        selected_date = selected_datetime.date()
        dias_semana = [
            "Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
            "Sexta-feira", "Sábado", "Domingo"
        ]
        dia_semana_idx = selected_date.dayOfWeek() - 1
        return dias_semana[dia_semana_idx]

    def update_dia_semana(self, selected_datetime):
        dia_semana_sem_data = self.get_dia_semana_text(selected_datetime)
        self.dia_semana.setText(dia_semana_sem_data)

    def handle_localizacao_change(self):
        self.update_fields_based_on_tecnico()
        selected_localizacao = self.combo_box_localizacao.currentText()
        if selected_localizacao in ["Folga", "Férias"]:
            self.combo_box_unidade.setDisabled(True)
            self.combo_box_unidade.setStyleSheet(
                self.combo_box_unidade.styleSheet() + "background-color: #FFD700;"
            )
        else:
            self.combo_box_unidade.setDisabled(False)
            self.combo_box_unidade.setStyleSheet(
                self.combo_box_unidade.styleSheet().replace("background-color: #FFD700;", "") +
                "background-color: #fff;"
            )

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            self.handle_enter_press()
        else:
            super().keyPressEvent(event)

    def send_emails(self):
        """Método de envio rápido de e-mail (sem abrir a consulta)."""
        tecnicos = list(self.technician_schedules.keys())
        dialog = EmailSelectionDialog(tecnicos, self.periodo_inicio, self.periodo_fim)
        if dialog.exec_() == QDialog.Accepted:
            typed_text = dialog.selected_tecnicos[0]
            periodo_inicio = dialog.periodo_inicio
            periodo_fim = dialog.periodo_fim

            try:
                df_existing = pd.read_excel(self.planilha_path)
                df_existing.columns = df_existing.columns.str.upper()
            except FileNotFoundError:
                QMessageBox.warning(self, "Erro", "A planilha selecionada não foi encontrada.")
                return

            if df_existing.empty:
                QMessageBox.information(self, "Aviso", "Não há dados na planilha para enviar.")
                return

            df_existing['DATA/HORA INICIO'] = pd.to_datetime(
                df_existing['DATA/HORA INICIO'], format="%d/%m/%Y %H:%M:%S", errors='coerce'
            )
            start_date = QDateTime(periodo_inicio, QTime(0, 0)).toPyDateTime()
            end_date = QDateTime(periodo_fim, QTime(23, 59, 59)).toPyDateTime()

            df_filtered = df_existing[
                (df_existing['DATA/HORA INICIO'] >= start_date) &
                (df_existing['DATA/HORA INICIO'] <= end_date) &
                (df_existing['TÉCNICO'].str.contains(typed_text, case=False, na=False))
            ]

            if df_filtered.empty:
                QMessageBox.information(self, "Aviso", "Não há registros para o técnico e período selecionados.")
                return

            try:
                locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
            except:
                locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')

            grouped = df_filtered.groupby('TÉCNICO')
            for tecnico, group in grouped:
                email = technician_emails.get(tecnico)
                if not email:
                    QMessageBox.warning(self, "Aviso", f"E-mail do técnico {tecnico} não encontrado.")
                    continue
                message = f"""
                <html>
                    <body>
                        <p>Prezado(a) <b>{tecnico}</b>,</p>
                        <p>Segue sua escala:</p>
                        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
                            <tr>
                                <th>DIA DA SEMANA</th>
                                <th>DATA</th>
                                <th>UNIDADE</th>
                                <th>LOCALIZAÇÃO</th>
                                <th>JUSTIFICATIVA</th>
                                <th>CARD</th>
                            </tr>
                """
                for idx, row in group.iterrows():
                    dia_semana = row['DIA DA SEMANA'] if pd.notnull(row['DIA DA SEMANA']) else ''
                    data_hora_inicio = row['DATA/HORA INICIO']
                    data = data_hora_inicio.strftime('%d/%m/%Y') if pd.notnull(data_hora_inicio) else ''
                    unidade = row['UNIDADE'] if pd.notnull(row['UNIDADE']) else ''
                    localizacao = row['LOCALIZAÇÃO'] if pd.notnull(row['LOCALIZAÇÃO']) else ''
                    justificativa = row['JUSTIFICATIVA'] if pd.notnull(row['JUSTIFICATIVA']) else ''
                    card = row['CARD'] if pd.notnull(row['CARD']) else ''

                    message += f"""
                            <tr>
                                <td>{dia_semana}</td>
                                <td>{data}</td>
                                <td>{unidade}</td>
                                <td>{localizacao}</td>
                                <td>{justificativa}</td>
                                <td>{card}</td>
                            </tr>
                    """
                message += """
                        </table>
                        <br>
                        <img src="cid:MinhaImagem" alt="Assinatura" />
                    </body>
                </html>
                """
                send_email(email, "Escala Semanal", message)

            reply = QMessageBox.question(
                self,
                'Enviar para Gestores',
                'Deseja enviar e-mails para os gestores das unidades?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                df_units = df_filtered.copy()
                df_units['DATA'] = df_units['DATA/HORA INICIO'].dt.date
                unit_grouped = df_units.groupby(['UNIDADE', 'DATA'])
                for (unidade, data_visita), group in unit_grouped:
                    gestor_email = unit_manager_emails.get(unidade)
                    if not gestor_email:
                        QMessageBox.warning(self, "Aviso", f"E-mail do gestor da unidade {unidade} não encontrado.")
                        continue

                    tecnicos_na_unidade = group['TÉCNICO'].unique()
                    tecnicos_lista = ', '.join(tecnicos_na_unidade)
                    data_inicio = group['DATA/HORA INICIO'].iloc[0]
                    horario_visita = data_inicio.strftime('%H:%M')

                    mensagem = f"""
                    <html>
                        <body>
                            <p>Prezado(a) Gestor(a)/PTA,<br><br>
                            Hoje foi agendada a visita do técnico <b>{tecnicos_lista}</b> da Liberty Health
                            para <b>apoio e acompanhamento na utilização do SGHX</b>
                            na unidade <b>{unidade}</b> a partir das {horario_visita}
                            do dia {data_visita.strftime('%d/%m/%Y')}.<br><br>
                            <br><br>
                            <img src="cid:MinhaImagem" alt="Assinatura" />
                            </p>
                        </body>
                    </html>
                    """
                    data_envio = datetime.datetime.combine(data_visita, datetime.time(4, 0))
                    try:
                        send_email(gestor_email, f"Visita Técnica - {data_visita.strftime('%d/%m/%Y')}", mensagem, send_time=data_envio)
                    except Exception as e:
                        QMessageBox.warning(self, "Erro", f"Falha ao agendar e-mail para {gestor_email}: {e}")

                QMessageBox.information(self, "Sucesso", "E-mails enviados aos técnicos e gestores das unidades.")
            else:
                QMessageBox.information(self, "Sucesso", "E-mails enviados apenas aos técnicos.")

# --------------------------------------------------
#               FUNÇÃO MAIN
# --------------------------------------------------
def main():
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('JT.ico'))

    selection_dialog = SelectionDialog()
    if selection_dialog.exec_() == QDialog.Accepted:
        planilha_path = selection_dialog.planilha_path
        periodo_inicio = selection_dialog.periodo_inicio
        periodo_fim = selection_dialog.periodo_fim
        choice = selection_dialog.choice

        form = ScheduleForm(planilha_path, periodo_inicio, periodo_fim)

        # Se escolher "Consultar Escala", ignoramos o período inicial
        # e exibimos tudo no ConsultaEscalaDialog
        if choice == 'consultar_escala':
            form.consultar_escala()

        sys.exit(app.exec_())
    else:
        sys.exit()

if __name__ == '__main__':
    main()